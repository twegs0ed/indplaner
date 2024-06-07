from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font
import io
import os
from order.models import Order, Firm
import qrcode
import openpyxl
import uuid
from tools.models import Tools, Toolsonwarehouse, Priem
from django.views.generic import ListView
from datetime import datetime, timedelta
import zipfile
from io import StringIO
import re
#from django.contrib.sites.models import Site

# Create your views here.
def count_c(det,det_in_zag, lenght, ws):
    if det_in_zag == 1:
        return {'ws' : ws, 'count' : det}
    
    
    if det_in_zag == None:
        det_in_zag=1

    if det is None or lenght is None:
        return {'ws' : ws, 'count' : ''}
    
    
    try:
        lenght=int(lenght)
    except:
        return {'ws' : ws, 'count' : ''}
    if not det_in_zag:det_in_zag=1
    if det/det_in_zag < 1:
        lenght=((lenght-30)/det_in_zag)*det+30
        ws.cell(row=4, column=7).value= str(int(lenght))

        return {'ws': ws, 'count' : 1}
    if det/det_in_zag > 1:
        if det%det_in_zag != 0:
            det+=1
        count = int((det/det_in_zag)+0.9)
        l_z =(lenght - 30)/det_in_zag
        lenght = ((det*l_z) + (30*count))/count
        #count=(lenght/det)*(det//det_in_zag)+((lenght-30)/det_in_zag)*det%det_in_zag+30
        ws.cell(row=4, column=7).value= str(int(lenght))
        return {'ws' : ws, 'count' : count}

    
    l_zag = lenght/ det
    l_fin = l_zag * det
    return {'ws' : ws, 'count' : float(det/det_in_zag)}
def printmk(request, id):
    first_name = [x+'.' for x in request.user.first_name if x.isupper()]
    fio = request.user.last_name+' '+''.join(first_name)
    order=Order.objects.get(pk=id)
    wb = load_workbook(filename = 'static/xls/mk.xlsx')
    ws = wb.active
    ws.cell(row=21, column=8).value=str(fio)

    if order.firm != None:
        ws.cell(row=2, column=2).value=str(order.firm)
    else:
        ws.cell(row=2, column=2).value=' '
    ws.cell(row=3, column=2).value=str(order.tool)
    if order.exp_date != None:
        ws.cell(row=4, column=2).value= format(order.exp_date, '%d.%m.%Y')
    else:
        ws.cell(row=4, column=2).value= ' '
    if order.tool.material_n != None:
        ws.cell(row=3, column=7).value= str(order.tool.material_n)
    else:
        ws.cell(row=3, column=7).value= ' '
    if order.tool.stock_sizes != None:
        ws.cell(row=4, column=7).value= str(order.tool.stock_sizes)
    else:
        ws.cell(row=4, column=7).value= ' '
    if order.tool.count_in_one_stock != None:
        count_cc = count_c(int(float(order.count)),int(float(order.tool.count_in_one_stock)), order.tool.stock_sizes, ws)
        ws=count_cc['ws']
        ws.cell(row=5, column=7).value= count_cc['count']
    else:
        ws.cell(row=5, column=7).value= ' '


    ws.cell(row=5, column=2).value=str(order.count)
    name = str(uuid.uuid4())
    
    image = qrcode.make(request.META['HTTP_HOST']+'/work/work/add/'+str(order.tool.id))
    
    image.save('static/xls/'+name+'.gif')# Напишите здесь свой код :-)


    img = openpyxl.drawing.image.Image('static/xls/'+name+'.gif')
    img.height = 150
    img.width = 150
    img.anchor = 'A24' # Or whatever cell location you want to use.
    ws.add_image(img)
    
   
    #ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    wb.save('static/xls/'+name+'.xlsx')

    file = io.open('static/xls/'+name+'.xlsx', "rb", buffering = 1024*256)
    file.seek(0)
    response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename= '+name+'.xlsx'
    
    file.close()
    if os.path.isfile('static/xls/'+name+'.xlsx'): os.remove('static/xls/'+name+'.xlsx')
    if os.path.isfile('static/xls/'+name+'.gif'): os.remove('static/xls/'+name+'.gif')

    return response

def getouts(request, firm):
    orders = Order.objects.filter(firm=firm).all()
    firm_obj = Firm.objects.get(id=firm)
    tools=[]
    '''for order in orders:
        if order not in temp:
            temp.append(order)
    orders=temp
    print(orders)'''
    for order in orders:
        t=Tools.objects.filter(tool=order.tool).filter(giveout_date__gte=firm_obj.date).all()
        if t:
            for t_c in t:
                t_c.tool = order.tool
                t_c.order_count=order.count
                t_c.order_id=order.id
                tools.append(t_c)
        else:
            t_c=Tools()
            t_c.tool = order.tool
            t_c.order_count=order.count
            t_c.order_id=order.id
            t_c.giveout_date = '<p style="background-color: #FF1820; color: #293133">не выдавалось</p>'
            tools.append(t_c)
        
    temp=[]
    for tool in tools:
        if tool not in temp:
            temp.append(tool)
    tools=temp
    tmp = {'tools':tools, 'firm':firm_obj}
    return render(request, 'getouts.html', tmp )



def priems(request, firm):
    orders = Order.objects.filter(firm=firm).all()
    firm_obj = Firm.objects.get(id=firm)
    tools=[]
    '''for order in orders:
        if order not in temp:
            temp.append(order)
    orders=temp
    print(orders)'''
    for order in orders:
        t=Priem.objects.filter(tool=order.tool).filter(giveout_date__gte=firm_obj.date).all()
        if t:
            for t_c in t:
                t_c.tool = order.tool
                t_c.order_count=order.count
                t_c.order_id=order.id
                tools.append(t_c)
        else:
            t_c=Tools()
            t_c.tool = order.tool
            t_c.order_count=order.count
            t_c.order_id=order.id
            t_c.giveout_date = '<p style="background-color: #FF1820; color: #293133">не принималось</p>'
            tools.append(t_c)
        
    temp=[]
    for tool in tools:
        if tool not in temp:
            temp.append(tool)
    tools=temp
    tmp = {'tools':tools, 'firm':firm_obj}
    return render(request, 'priems.html', tmp )

def printmkall(request, id):


    first_name = [x+'.' for x in request.user.first_name if x.isupper()]
    fio = request.user.last_name+' '+''.join(first_name)


    firm = Firm.objects.get(pk=id)
    orders=Order.objects.filter(firm = firm)
    
    files=[]
    gifs=[]
    for order in orders:
        wb = load_workbook(filename = 'static/xls/mk.xlsx')
        ws = wb.active
        ws.cell(row=21, column=8).value=str(fio)

        if order.firm != None:
            ws.cell(row=2, column=2).value=str(order.firm)
        else:
            ws.cell(row=2, column=2).value=' '
        ws.cell(row=3, column=2).value=str(order.tool)
        if order.exp_date != None:
            ws.cell(row=4, column=2).value= format(order.exp_date, '%d.%m.%Y')
        else:
            ws.cell(row=4, column=2).value= ' '
        if order.tool.material_n != None:
            ws.cell(row=3, column=7).value= str(order.tool.material_n)
        else:
            ws.cell(row=3, column=7).value= ' '
        if order.tool.stock_sizes != None:
            ws.cell(row=4, column=7).value= str(order.tool.stock_sizes)
        else:
            ws.cell(row=4, column=7).value= ' '
        if order.tool.count_in_one_stock != None:
            count_cc = count_c(int(float(order.count)),int(float(order.tool.count_in_one_stock)), order.tool.stock_sizes, ws)
            ws=count_cc['ws']
            ws.cell(row=5, column=7).value= count_cc['count']
        else:
            ws.cell(row=5, column=7).value= ' '


        ws.cell(row=5, column=2).value=str(order.count)
        name = re.sub(r'[^а-яА-Яa-zA-Z0-9.\s]+', '', order.tool.title)
        
        image = qrcode.make(request.META['HTTP_HOST']+'/work/work/add/'+str(order.tool.id))
        
        image.save('static/xls/'+name+'.gif')# Напишите здесь свой код :-)


        img = openpyxl.drawing.image.Image('static/xls/'+name+'.gif')
        img.height = 150
        img.width = 150
        img.anchor = 'A24' # Or whatever cell location you want to use.
        ws.add_image(img)
        
    
        #ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        wb.save('static/xls/'+name+'.xlsx')

        file = io.open('static/xls/'+name+'.xlsx', "rb", buffering = 1024*256)
        file.seek(0)
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename= '+name+'.xlsx'
        
        file.close()
        files.append('static/xls/'+name+'.xlsx')
        gifs.append('static/xls/'+name+'.gif')

    comp_file = zipfile.ZipFile('static/xls/XLS.zip', 'w')
    for f in files:
        comp_file.write(f,os.path.basename(f))
    comp_file.close()
    file = io.open('static/xls/XLS.zip', "rb", buffering = 1024*256)
    file.seek(0)
    response = HttpResponse(file.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="XLS.zip"'
    
    file.close()
    for f in files:
        if os.path.isfile(f): os.remove(f)
    for f in gifs:
        if os.path.isfile(f): os.remove(f)
    if os.path.isfile('static/xls/XLS.zip'): os.remove('static/xls/XLS.zip')
    return response




def printpr(request, id):
    firm = Firm.objects.get(pk=id)
    orders=Order.objects.filter(firm = firm).all()
    assems=firm.assem.all()
    
    files=[]
    gifs=[]
    row=2
    wb = load_workbook(filename = 'static/xls/pr.xlsx')
    ws = wb.active
    
    for assem in assems:
        ws.cell(row=row, column=1).value=str(assem.title)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for tool in assem.tool.all():
            row+=1
            ws.cell(row=row, column=1).value=str(tool.title)
            order = Order.objects.filter(firm = firm).filter(tool = tool)
            n=0
            for or_c in order:
                print(orders)
                n+=or_c.count
                print(or_c.tool.title)
                print(or_c.count)
                print('!!!!!!!!!!!!')
                print(orders)
            
            else:
                count = 'error'
            ws.cell(row=row, column=2).value=str(n)
            
            pass
        row+=1
        pass
        row+=1


    name = re.sub(r'[^а-яА-Яa-zA-Z0-9.\s]+', '', firm.title)
    wb.save('static/xls/'+name+'.xlsx')

    file = io.open('static/xls/'+name+'.xlsx', "rb", buffering = 1024*256)
    file.seek(0)
    response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename= '+name+'.xlsx'
    
    file.close()
    if os.path.isfile('static/xls/'+name+'.xlsx'): os.remove('static/xls/'+name+'.xlsx')
    if os.path.isfile('static/xls/'+name+'.gif'): os.remove('static/xls/'+name+'.gif')

    return response
