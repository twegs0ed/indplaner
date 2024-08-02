from django.contrib import admin
from rangefilter.filters import DateRangeFilter, DateTimeRangeFilter
from .models import Toolsonwarehousezn, Toolszn, Priemzn, Workplace
from profiles.models import Profile
from order.models import Order, Firm
from import_export.admin import ExportActionMixin, ImportExportModelAdmin
from import_export import resources
from import_export.fields import Field
from django.utils.html import format_html
from import_export.widgets import ForeignKeyWidget, CharWidget, ManyToManyWidget
from django.db import models
from django.forms import TextInput, Textarea
import re
from material.models import Material
from django.shortcuts import render
import datetime
from work.models import Work
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import openpyxl
import uuid
from io import StringIO
import io
from django.http import HttpResponse
import os
#import copy


def order_it(modeladmin, request, queryset):
    for obj in queryset:
        i=Order.get_count_ordered(obj)
        ord=Order()
        '''c_ord=get_object(Order,obj)
        if c_ord:
            c_ord.delete()'''
        ord.tool=obj
        ord.worker=Profile.objects.get(pk=10)
        ord.count=obj.min_count-obj.count-i
        ord.save()

def get_object(self, obj):
    try:
        return Order.objects.filter(tool=obj)
    except Order.DoesNotExist:
        return False


order_it.short_description = "В заявку"

class ToolsznResource(resources.ModelResource):
    
   

    class Meta:
        model = Toolszn
        
        fields = ('tool', 'count','worker__bio', 'giveout_date')
        export_order = ('tool', 'worker__bio', 'count')

class ForeignKeyWidgetWithCreation (ForeignKeyWidget):

    def clean(self, value, row=None, *args, **kwargs):
        try:
            return super(ForeignKeyWidgetWithCreation, self).clean(value, row, *args, **kwargs)
        except:
            return self.model.objects.create(**{self.field: value})
        
class ToolsonwarehouseznResource(resources.ModelResource):
    material_n = Field(column_name='material_n',attribute='material_n',widget=ForeignKeyWidgetWithCreation(model=Material, field='title'))
    formfield_overrides = {
        models.TextField: {'widget': Textarea(attrs={'rows':5, 'cols':40})},
    }
    #tool = Field(attribute='title', column_name='title', widget=CharWidget(),)
    similar = Field()
    workplace = Field(
        column_name='workplace',
        attribute='workplace',
        widget=ForeignKeyWidgetWithCreation(model=Workplace, field='name'))
    
    class Meta:
        model = Toolsonwarehousezn
        #skip_unchanged=False
        fields = ('title', 'count', 'workplace__name','material_n' ,'similar','stock_sizes', 'count_in_one_stock', 'cover', 
                  'norm_lentopil_p','norm_lentopil','norm_plazma_p','norm_plazma','norm_turn_p','norm_turn','norm_mill_p',
                  'norm_mill','norm_turnun_p','norm_turnun','norm_millun_p',
                  'norm_millun','norm_electro_p','norm_electro','norm_slesarn','norm_sverliln_p','norm_sverliln','norm_rastoch_p','norm_rastoch')
        exclude = ('id',)
        export_order = ('title','count', 'workplace')
        import_id_fields=['title']
        #import_id_field = 'title'
    def dehydrate_similar(self, tool): 
        res=''
        if tool.id:
            if tool.similar:
                for t in tool.similar.all():
                    res+=t.title+', '
        return res

    
        
class ToolsonwarehouseznAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    
    formfield_overrides = {models.TextField: {'widget': Textarea(attrs={'rows':5, 'cols':40})},}
    resource_class = ToolsonwarehouseznResource
    autocomplete_fields = []
    list_display = ('tool', 'count', 'workplace' )
    list_filter = ('workplace',)
    search_fields = ['tool__title']
    ordering = ['tool']
    list_editable = []
    
    def get_search_results(self, request, queryset, search_term):
        search_term=str(search_term).upper()
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        return queryset, use_distinct
    def firms(self, obj):
        orders=Order.objects.filter(tool=obj)
        t=''
        f=[]
        for o in orders:
            if o.firm:
                f.append(o.firm.title)
            pass
        for x in f: 
            if f.count(x) > 1: 
                f.remove(x) 
        for x in f:
            t+=' <a href = "/order/order/?q='+x+'">'+x+'</a> '+'</br>'
            #t+=' <a href = "/tools/toolsonwarehouse/?q='+str(f.tool.title)+'">'+str(f.title)+'</a> '
        return format_html(t)
    firms.short_description = "Изделия"
    
    def similar_c(self,obj):
        res=''
        if obj.similar:
            for t in obj.similar.all():
                
                res+=' <a href = "/tools/toolsonwarehouse/'+str(t.id)+'">'+t.title+'</a> '+'</br>'+', '
        return format_html(res)
    similar_c.short_description = "Похожие"
    class Media:
        js = [
                'js/list_filter_collapse.js'
                ]
def m15_action(modeladmin, request, queryset):
    tools = list(queryset)
    first_name = [x+'.' for x in request.user.first_name if x.isupper()]
    fio = request.user.last_name+' '+''.join(first_name)
    wb = load_workbook(filename = 'static/xls/m15.xlsx')
    ws = wb.active
    
    ws.cell(row=9, column=4).value=str(datetime.datetime.now().date().strftime("%d.%m.%Y"))
    row=17
    ws.row_dimensions[row].height = 12
    thins = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")
    #ws['B'+str(row)].border = Border(top=thins, bottom=thins, left=thins, right=thins) 
    for t in tools:
        ws.insert_rows(row)
        ws['B'+str(row)].border = Border(top=thins, bottom=thins, left=thins, right=thins) 
        ws['C'+str(row)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        ws['D'+str(row)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        ws.row_dimensions[row+1].height = 12
        ws.merge_cells(start_row=row+1, start_column=4, end_row=row+1, end_column=7)
        ws.cell(row=row, column=4).value=str(t.tool.tool.title)
        ws.cell(row=row, column=11).value=str(t.count)
        ws.cell(row=row, column=12).value=str(t.count)
        row+=1
        
        pass
    print(row)
    for col in ws['B17:S'+str(row)]:
            for cell in col:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin) 
    ws.delete_rows(row)
    
    

   
    name = str(uuid.uuid4())
   
    

    #ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    wb.save('static/xls/'+name+'.xlsx')

    file = io.open('static/xls/'+name+'.xlsx', "rb", buffering = 1024*256)
    file.seek(0)
    response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename= '+name+'.xlsx'
    
    file.close()
    if os.path.isfile('static/xls/'+name+'.xlsx'): os.remove('static/xls/'+name+'.xlsx')
    return response
m15_action.short_description = "M15"#заяввка

class ToolsznAdmin(ExportActionMixin, admin.ModelAdmin):
    actions = [m15_action]
    formfield_overrides = {
        models.TextField: {'widget': Textarea(attrs={'rows':5, 'cols':40})},
    }
    #raw_id_fields = ('worker', 'tool')
    autocomplete_fields = ['tool','worker']
    resource_class = ToolsznResource
    list_display = ('tool', 'worker', 'count', 'giveout_date', 'text')
    list_filter = (('giveout_date', DateRangeFilter), 'worker')
    search_fields = ['tool__tool__title', 'worker__bio']
    ordering = ['-giveout_date']
    list_editable = ['text']
    def get_search_results(self, request, queryset, search_term):
        #search_term=re.sub("[^\d\.]", "", str(search_term))
        search_term=str(search_term).upper()
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        #queryset |= self.model.objects.filter(tool__title=search_term)
        return queryset, use_distinct
    class Media:
        js = [
                'js/list_filter_collapse.js'
                ]
    pass


  
class Rec_ToolsAdmin(ExportActionMixin, admin.ModelAdmin):
    #raw_id_fields = ('worker', 'tool')
    autocomplete_fields = ['worker', 'tool']
    resource_class = ToolsznResource
    list_display = ('worker', 'tool', 'count', 'giveout_date')
    list_filter = (('giveout_date', DateRangeFilter), 'worker', 'tool')
    search_fields = ['tool__title', 'worker__bio']
    ordering = ['-giveout_date']
    pass

class PriemznResource(resources.ModelResource):
    firms = Field()
    tool = Field(
        column_name='tool',
        attribute='tool',
        widget=ForeignKeyWidgetWithCreation(model=Toolsonwarehousezn, field='title'))
    #firm = Field(
        #column_name='firm',
        #attribute='firm',
        #widget=ForeignKeyWidgetWithCreation(model=Firm, field='title'))
    class Meta:
        model = Priemzn

        fields = ('tool', 'count','place','giveout_date','text', 'firms')
        export_order = ('tool', 'count')
        exclude = ('id',)
        import_id_fields = ('tool', 'count')
    def dehydrate_firms(self, obj):
        orders=Order.objects.filter(tool=obj.tool).order_by('-order_date_worker')[:15]
        t=''
        f=[]
        for o in orders:
            if o.firm:
                f.append(o.firm)
            pass
        for x in f: 
            if f.count(x) > 1: 
                f.remove(x) 
        for x in f:
            if x:
                t+=x.title+' - '+str(x.count)+' шт.'+'\n'
            #t+=' <a href = "/tools/toolsonwarehouse/?q='+str(f.tool.title)+'">'+str(f.title)+'</a> '
        return format_html(t)
    dehydrate_firms.short_description = "Изделия"
def svod_action(modeladmin, request, queryset):
    orders=[]
    firms = []
    #priems = queryset
    priems = list(queryset)
    for priem in priems:
        for pr in priems:
            if pr.tool==priem.tool and pr.giveout_date!=priem.giveout_date:
                priem.count+=pr.count
                priems.remove(pr)
    for priem in priems:
        for order_c in Order.objects.filter(tool=priem.tool).all():
            priem.c_count=priem.count
            orders.append([order_c, priem])
            if order_c.firm not in firms:
                if not order_c.firm:
                    break
                if not order_c.firm.exp_date:
                    order_c.firm.exp_date = datetime.date.today()
                firms.append(order_c.firm)


    firms = get_firms_for_priem(request, firms)
    orders_c = get_result_for_period(firms, priems)
    

    '''i=0
    for order in orders:
        print(order[0].tool,'-',order[0].count,'заказ')
        print(order[1].tool,'-',order[1].count,'прием')
        i+=1
        if i==5: break'''
        

    return render(request, 'priem.html', {'orders_c': orders_c,'orders': orders,'firms': firms,'title':u'Прием на склад'})
svod_action.short_description = "Сводная"#заяввка

def get_firms_for_priem(request, firms):
    start_date=datetime.datetime.strptime(request.GET.get('giveout_date__range__gte'), '%d.%m.%Y').date()
    end_date=datetime.datetime.strptime(request.GET.get('giveout_date__range__lte'), '%d.%m.%Y').date()
    firms.sort(key=lambda x: x.exp_date)
    firms = list(filter(lambda x: x.exp_date < (end_date+datetime.timedelta(40)), firms))
    firms = list(filter(lambda x: x.exp_date > (start_date-datetime.timedelta(40)), firms))
    return firms
def get_result_for_period(firms, priems):
    orders = []
    for firm in firms:
        for order in Order.objects.filter(firm=firm).all():
            orders.append(order)
            pass
    priems=list(priems)
    for priem in priems.copy():
        while priem.count > 0:
            for order in orders:
                if order.tool == priem.tool:
                    if order.count < priem.count:
                        order.percent = 100
                        priem.count-=order.count
                    elif order.count == priem.count:
                        order.percent = 100
                        priem.count=0
                        priems.remove(priem)
                        break
                    elif order.count > priem.count:
                        order.percent = priem.count/order.count*100
                        priem.count=0
                        priems.remove(priem)
                        break
                    #print(order.tool,order.percent, order.firm)
            
            break
        
    
    return orders

class PriemznAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    actions = [svod_action]
    resource_class = PriemznResource
    formfield_overrides = {
        models.TextField: {'widget': Textarea(attrs={'rows':5, 'cols':40})},
    } 
    autocomplete_fields = ['tool', 'place','worker']
    list_display = ('tool', 'count', 'giveout_date', 'text', 'firms')
    #list_filter = (('giveout_date', DateRangeFilter))
    list_filter = (('giveout_date', DateRangeFilter),)
    search_fields = ['tool__title']
    list_editable = ['text']
    def get_search_results(self, request, queryset, search_term):
        search_term=str(search_term).upper()
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        
        try:
            search_term_as_int = int(search_term)
        except ValueError:
            pass
        else:
            queryset |= self.model.objects.filter(tool=search_term_as_int)
        return queryset, use_distinct
    pass
    def firms(self, obj):
        orders=Order.objects.filter(tool=obj.tool)
        t=''
        f=[]
        for o in orders:
            if o.firm:
                f.append(o.firm.title)
            pass
        for x in f: 
            if f.count(x) > 1: 
                f.remove(x) 
        for x in f:
            t+=' <a href = "/order/order/?q='+x+'">'+x+'</a> '+'</br>'
            #t+=' <a href = "/tools/toolsonwarehouse/?q='+str(f.tool.title)+'">'+str(f.title)+'</a> '
        return format_html(t)
    firms.short_description = "Изделия"
    class Media:
        js = [
                'js/list_filter_collapse.js'
                ]




admin.site.register(Toolszn, ToolsznAdmin)
admin.site.register(Priemzn, PriemznAdmin)
admin.site.register(Toolsonwarehousezn, ToolsonwarehouseznAdmin)
#admin.site.register(Rec_Tools, Rec_ToolsAdmin)
